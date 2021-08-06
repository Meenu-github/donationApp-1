import streamlit as st
import openpyxl as pxl
import base64
wb = pxl.load_workbook('BloodDonation.xlsx')
ws = wb.active
maxrow= ws.max_row+1

def bloodDonate() :
    main_bg = "book1.gif"
    main_bg_ext = "gif"
    st.markdown(f"""
    <style>
    .reportview-container {{
        background: url(data:image/{main_bg_ext};base64,{base64.b64encode(open(main_bg, "rb").read()).decode()})
        background-position: center; background-size: cover; background-repeat: no repeat;
    }}
    </style>
    """,
    unsafe_allow_html=True
)
    st.title("Blood Donation")
    st.write("This is the blood donation page.")
    st.write("Here if you are willing to donate blood\n you have to register yourself.")
    
    with st.form(key="Registration for Blood Donation"):
        bloodname = st.text_input("Enter your name : ")
        bgrp = st.text_input("Enter your blood group : ")
        age = st.slider(label="Enter your age ", min_value=18, max_value=45)
        blood_phone =  st.text_input("Enter your Phone number : ")
        date = st.date_input(label="Date")
        
        submissionblood = st.form_submit_button(label="Submit")
        if submissionblood==True:
            
            
            ws.cell(row=maxrow,column=1).value = bloodname
            ws.cell(row=maxrow,column=2).value = bgrp
            ws.cell(row=maxrow,column=3).value = age
            ws.cell(row=maxrow,column=4).value = blood_phone
            ws.cell(row=maxrow,column=5).value = date
            
            wb.save('BloodDonation.xlsx')
            
            st.success("Successfully submitted your data. Thanks for registering.")


    



    #adding date picker..

