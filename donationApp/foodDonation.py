import streamlit as st
import openpyxl as pxl
wb = pxl.load_workbook('FoodDonation.xlsx')
ws = wb.active
maxrow= ws.max_row+1
#ws.title = "fooddonation"


def foodDonate() :
    st.markdown("FOOD DONATION")
    st.title("Welcome to the food donation page, your donated food can bring hope in someones life of survival.\nCome let us donate food for needy one.\nYou don't have to walk and donate it you just have to register yourself and we will pick the food from your house address that will be provided.")
    st.write("Here if you are willing to donate food\n you have to register yourself.")
    with st.form(key="Register for food donation"):
        namefood = st.text_input("Enter your name : ")
        foodaddress = st.text_input("Enter your address please : ")
        food_phone = st.text_input("Enter your phone  Number : ")
        
        foodsubmission = st.form_submit_button(label="Submit")
        if foodsubmission==True:
            ws.cell(row=maxrow,column=1).value = namefood
            ws.cell(row=maxrow,column=2).value = foodaddress
            ws.cell(row=maxrow,column=3).value = food_phone
            
            wb.save('FoodDonation.xlsx')
            st.success("Successfully registered for food donation")
        else:
            st.info("Please submit the form.")

        

