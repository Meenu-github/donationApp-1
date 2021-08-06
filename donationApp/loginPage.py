import streamlit as st
import openpyxl as pxl
import pandas as pd
import base64

#import gspread
#from gspread_dataframe import get_as_dataframe 
#proj = pd.read_excel('proj.xlsx')
#exp = pxl.load_workbook('proj.xlsx')

projDemo = pd.read_excel('proj.xlsx')
exp = pxl.load_workbook('proj.xlsx')
sheet = exp.active
maxrow = sheet.max_row+1
#sheet = exp['Sheet2']

def type(selectRole):
    if selectRole == 'Hospital':
        exp = pd.read_excel('BloodDonation.xlsx')
        #new = pxl.load_workbook('../donationApp/BloodDonation.xlsx')
        #newsheet = new.active
        #sheet = exp.active
        #maxrow = sheet.max_row+1
        st.table(exp)
        

    if selectRole == 'Food Distributor':
        exp1 = pd.read_excel('FoodDonation.xlsx')
        #new = pxl.load_workbook('../donationApp/FoodDonation.xlsx')
        #newsheet = new.active
        #sheet = exp1.active
        #maxrow = sheet.max_row+1
        st.table(exp1)
        

    if selectRole == 'Orphanage':
        exp2 = pd.read_excel('BookDonation.xlsx')
        #new = pxl.load_workbook('../donationApp/BookDonation.xlsx')
        #newsheet = new.active
        #sheet = exp2.active
        #maxrow = sheet.max_row+1
        st.table(exp2)
        


def loginPages():
    main_bg = "book1.gif"
    main_bg_ext = "gif"
    st.markdown(f"""
    <style>
    .reportview-container {{
        background: url(data:image/{main_bg_ext};base64,{base64.b64encode(open(main_bg, "rb").read()).decode()})
        
    }}
    </style>
    """,
    unsafe_allow_html=True
)
    page_name = ['Register', 'Login']
    page = st.radio('Choose Register/Login', page_name)
    if page == 'Register':
        with st.form(key="Sign up"):
            userName = st.text_input("Username")
            email = st.text_input("E-mail Id")
            password = st.text_input("Password", type="password")
            conf_password = st.text_input('Confirm your password', type="password")
            role = ['Hospital', 'Food Distributor', 'Orphanage']
            selectRole = st.selectbox("Role", role)
            if password == conf_password:
                pass
            else:
                st.warning("password do not match with confirm password")
               

            sheet.cell(row=maxrow, column=1).value = userName
            sheet.cell(row=maxrow, column=2).value = email
            sheet.cell(row=maxrow, column=3).value = password
            sheet.cell(row=maxrow, column=4).value = selectRole

            submissionButton = st.form_submit_button(label="Sign up")
            if submissionButton == True:
                exp.save('proj.xlsx')
                st.success("Successfully sign up")
                #data = ['Yes, I want donor data', 'No, I don\'t want donor data']
                #b2 = st.selectbox('Do you want data of book donor :', data)
                #if b2 == 'Yes, I want donor data':
                st.info('Giving data of donor')

                type(selectRole)
                #if b2 == 'No, I don\'t want donor data':
                    #st.warning('YOU SELECTED NO.')

                

    if page == 'Login' :
        with st.form(key="Login"):

            userName = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submissionButton = st.form_submit_button(label="Login")
            if submissionButton == True:

                for i in range(2, sheet.max_row):
                    if((sheet.cell(row=i, column=2).value == userName)):
                        if((sheet.cell(row=i, column=3).value == password)):
                            y = sheet.cell(row=i, column=4).value
                            st.success('Successfully Login')
                            #data = ['Yes, I want donor data', 'No, I don\'t want donor data']
                            #b2 = st.selectbox('Do you want data of book donor :', data)
                            #if b2 == 'Yes, I want donor data':
                            st.info('Giving data of donor')
                            type(y)
                            #if b2 == 'No, I don\'t want donor data':
                                #st.warning('YOU SELECTED NO.')
                            
                            

                        else:
                            st.error(
                                "either username or password is incorrect")




#st.table(projDemo)

# for i in range(1, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         st.write(sheet.cell(row=i, column=j).value, end="  ")
