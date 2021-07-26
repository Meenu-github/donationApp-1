import streamlit as st
import openpyxl as pxl
wb = pxl.load_workbook('BookDonation.xlsx')
ws = wb.active
maxrow= ws.max_row+1

def bookdonate():
    st.markdown("BOOK DONATION")
    st.title("Welcome to the book donation page, your old book can bring light in someones future.\nCome let us donate books for needy one.\nYou don't have to walk and donate it you just have to register yourself and we will pick the book from your house address that will be provided.")
    st.write("Here if you are willing to donate book\n you have to register yourself.")
    
    with st.form(key="Registration for Blood Donation"):
    

        bookname = st.text_input("Enter your name : ")
        address = st.text_input("Enter your address please : ")
        book_phone = st.text_input("Enter your phone Number : ")
        
        submissionbook = st.form_submit_button(label="Submit")
        if submissionbook==True:
            ws.cell(row=maxrow,column=1).value = bookname
            ws.cell(row=maxrow,column=2).value = address
            ws.cell(row=maxrow,column=3).value = book_phone
            
            
            wb.save('BookDonation.xlsx')
            
            st.success("Successfully submitted the form.")
        else:
            st.info("Please submit the form.")

